#!/usr/bin/env python

import os
import sys
import urllib2
import zipfile
import numpy as np
import pandas as pd
import IPython as ip
from openpyxl import load_workbook
from misc import intersect_index


def main():

    # Read URLs to GDSC datasets
    urls_file = sys.argv[1]
    urls = []
    with open(urls_file) as f:
        for line in f:
            if line.startswith('http://') or line.startswith('https://'):
                urls.append(line[:-1])


    # Create data folder
    directory = '%s/data' % os.getcwd()
    if not os.path.exists(directory):
        os.makedirs(directory)


    # Download datasets
    for url in urls:
        print 'Downloading %s' % url
        local_fn = os.path.join(directory, os.path.basename(url))
        remote_file = urllib2.urlopen(url)
        with open(local_fn, 'wb') as local_file:
            local_file.write(remote_file.read())
        remote_file.close()
        if local_fn.endswith('.zip'):
            with zipfile.ZipFile(local_fn, 'r') as zip_ref:
                zip_ref.extractall(directory)

    print 'Preprocessing the GDSC dataset...'

    # Read Gene expression dataset
    GEX_file = '%s/Cell_line_RMA_proc_basalExp.txt' % directory
    GEX = pd.read_csv(GEX_file, sep='\t')
    GEX_gene_symbols = np.array(GEX['GENE_SYMBOLS'], dtype='str')
    GEX = GEX.drop(['GENE_SYMBOLS', 'GENE_title'], axis=1)
    GEX_cell_ids = np.array(GEX.columns, dtype='str')
    for i, cell_id in enumerate(GEX_cell_ids):
        GEX_cell_ids[i] = cell_id[5:]
    GEX = np.array(GEX.values, dtype=np.float).T

    # Read Exome sequencing dataset
    WES_file = '%s/CellLines_CG_BEMs/PANCAN_SEQ_BEM.txt' % directory
    WES = pd.read_csv(WES_file, sep='\t')
    WES_CG = np.array(WES['CG'], dtype='str')
    WES = WES.drop(['CG'], axis=1)
    WES_cell_ids = np.array(WES.columns, dtype='str')
    WES = np.array(WES.values, dtype=np.int).T

    # Read Copy number dataset
    CNV_file = '%s/CellLine_CNV_BEMs/PANCAN_CNA_BEM.rdata.txt' % directory
    CNV = pd.read_csv(CNV_file, sep='\t')
    CNV_cell_ids = np.array(CNV['Unnamed: 0'], dtype='str')
    CNV = CNV.drop(['Unnamed: 0'], axis=1)
    CNV_cna = np.array(CNV.columns, dtype='str')
    CNV = np.array(CNV.values, dtype=int)

    # Read Methylation dataset
    MET_file = '%s/METH_CELLLINES_BEMs/PANCAN.txt' % directory
    MET = pd.read_csv(MET_file, sep='\t')
    MET_met = np.array(MET['Unnamed: 0'], dtype='str')
    MET = MET.drop(['Unnamed: 0'], axis=1)
    MET_cell_ids = np.array(MET.columns, dtype='str')
    MET = np.array(MET.values, dtype=int).T

    # Read LOG_IC50 dataset
    IC50_file = '%s/TableS4A.xlsx' % directory
    wb = load_workbook(filename=IC50_file)
    sheet = wb['TableS4A-IC50s']
    IC50_cell_ids, IC50_cell_names = [], []
    for i in range(7, 997):
        IC50_cell_ids.append('%s' % sheet['A%s' % i].value)
        IC50_cell_names.append(('%s' % sheet['B%s' % i].value).strip())
    IC50_cell_ids = np.array(IC50_cell_ids, dtype='str')
    IC50_cell_names = np.array(IC50_cell_names, dtype='str')

    IC50_drug_ids, IC50_drug_names = [], []
    for i, (cell_row5, cell_row6) in enumerate(zip(sheet[5], sheet[6])):
        if i > 1:
            IC50_drug_ids.append('%s' % cell_row5.value)
            IC50_drug_names.append(('%s' % cell_row6.value).strip())
    IC50_drug_ids = np.array(IC50_drug_ids, dtype='str')
    IC50_drug_names = np.array(IC50_drug_names, dtype='str')

    IC50 = np.ones([IC50_cell_ids.shape[0], IC50_drug_ids.shape[0]]) * np.nan
    for i in range(7, 997):
        for j, cell in enumerate(sheet[i]):
            if j > 1:
                if cell.value != 'NA':
                    IC50[i - 7, j - 2] = cell.value

    # Read LOG_IC50 Threshold
    threshold_file = '%s/TableS5C.xlsx' % directory
    wb = load_workbook(filename=threshold_file)
    sheet = wb['Table-S5C binaryIC50s']
    threshold = []
    for i, cell in enumerate(sheet[7]):
        if i > 1:
            threshold.append(cell.value)
    threshold = np.array(threshold)
    drug_ids_file = '%s/TableS1F.xlsx' % directory
    wb = load_workbook(filename=drug_ids_file)
    sheet = wb['TableS1F_ScreenedCompounds']
    threshold_drug_ids = []
    for i in range(4, 269):
        threshold_drug_ids.append('%s' % sheet['B%s'%i].value)
    threshold_drug_ids = np.array(threshold_drug_ids)


    # Normalize IC50 by the threshold
    merged = intersect_index(IC50_drug_ids, threshold_drug_ids)
    IC50_keep_index = np.array(merged['index1'].values, dtype=np.int)
    IC50_drug_ids = IC50_drug_ids[IC50_keep_index]
    IC50 = IC50[:, IC50_keep_index]
    threshold_keep_index = np.array(merged['index2'].values, dtype=np.int)
    threshold_drug_ids = threshold_drug_ids[threshold_keep_index]
    threshold = threshold[threshold_keep_index]
    IC50_norm = - (IC50 - threshold)
    IC50_norm_min = np.min(IC50_norm[~np.isnan(IC50_norm)])
    IC50_norm = IC50_norm - IC50_norm_min


    # Save the GEX features and normalized IC50 dataset
    merged = intersect_index(GEX_cell_ids, IC50_cell_ids)
    GEX_keep_index = np.array(merged['index1'].values, dtype=np.int)
    IC50_keep_index = np.array(merged['index2'].values, dtype=np.int)
    GEX = GEX[GEX_keep_index]
    GEX_cell_ids = GEX_cell_ids[GEX_keep_index]
    GEX_cell_names = IC50_cell_names[IC50_keep_index]
    IC50 = IC50_norm[IC50_keep_index]
    np.savez('%s/GDSC_GEX.npz' % directory, X=GEX, Y=IC50, cell_ids=GEX_cell_ids, cell_names=GEX_cell_names,
             drug_ids=IC50_drug_ids, drug_names=IC50_drug_names, GEX_gene_symbols=GEX_gene_symbols)
    print 'Gene expression (GEX) dataset: {} cell lines, {} features, {} drugs'.format(GEX.shape[0], GEX.shape[1], IC50.shape[1])


    # Save the WES features and normalized IC50 dataset
    merged = intersect_index(WES_cell_ids, IC50_cell_ids)
    WES_keep_index = np.array(merged['index1'].values, dtype=np.int)
    IC50_keep_index = np.array(merged['index2'].values, dtype=np.int)
    WES = WES[WES_keep_index]
    WES_cell_ids = WES_cell_ids[WES_keep_index]
    WES_cell_names = IC50_cell_names[IC50_keep_index]
    IC50 = IC50_norm[IC50_keep_index]
    np.savez('%s/GDSC_WES.npz' % directory, X=WES, Y=IC50, cell_ids=WES_cell_ids, cell_names=WES_cell_names,
             drug_ids=IC50_drug_ids, drug_names=IC50_drug_names, WES_CG=WES_CG)
    print 'Whole-exome sequencing (WES) dataset: {} cell lines, {} features, {} drugs'.format(WES.shape[0], WES.shape[1], IC50.shape[1])


    # Save the CNV features and normalized IC50 dataset
    merged = intersect_index(CNV_cell_ids, IC50_cell_ids)
    CNV_keep_index = np.array(merged['index1'].values, dtype=np.int)
    IC50_keep_index = np.array(merged['index2'].values, dtype=np.int)
    CNV = CNV[CNV_keep_index]
    CNV_cell_ids = CNV_cell_ids[CNV_keep_index]
    CNV_cell_names = IC50_cell_names[IC50_keep_index]
    IC50 = IC50_norm[IC50_keep_index]
    np.savez('%s/GDSC_CNV.npz' % directory, X=CNV, Y=IC50, cell_ids=CNV_cell_ids, cell_names=CNV_cell_names,
             drug_ids=IC50_drug_ids, drug_names=IC50_drug_names, CNV_cna=CNV_cna)
    print 'Copy number variation (CNV) dataset: {} cell lines, {} features, {} drugs'.format(CNV.shape[0], CNV.shape[1], IC50.shape[1])


    # Save the MET features and normalized IC50 dataset
    merged = intersect_index(MET_cell_ids, IC50_cell_ids)
    MET_keep_index = np.array(merged['index1'].values, dtype=np.int)
    IC50_keep_index = np.array(merged['index2'].values, dtype=np.int)
    MET = MET[MET_keep_index]
    MET_cell_ids = MET_cell_ids[MET_keep_index]
    MET_cell_names = IC50_cell_names[IC50_keep_index]
    IC50 = IC50_norm[IC50_keep_index]
    np.savez('%s/GDSC_MET.npz' % directory, X=MET, Y=IC50, cell_ids=MET_cell_ids, cell_names=MET_cell_names,
             drug_ids=IC50_drug_ids, drug_names=IC50_drug_names, MET_met=MET_met)
    print 'Methylation (MET) dataset: {} cell lines, {} features, {} drugs'.format(MET.shape[0], MET.shape[1], IC50.shape[1])


    print 'Finished.'

if __name__ == '__main__':
    main()




